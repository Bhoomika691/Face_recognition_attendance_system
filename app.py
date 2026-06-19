import os, csv, threading, time, io
from datetime import datetime, timedelta
from pathlib import Path
import calendar

from flask import (Flask, render_template, Response, jsonify,
                   request, session, redirect, url_for, send_file)
from flask_cors import CORS
import cv2
import numpy as np
from deepface import DeepFace

# ── For PDF and Excel export ──────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__)
CORS(app)

# ── Secret key for session (change this to something random in production) ────
app.secret_key = "face_attendance_secret_2024"

# ── ADMIN CREDENTIALS (change these!) ────────────────────────────────────────
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# ── CONFIG ────────────────────────────────────────────────────────────────────
DUPLICATE_BLOCK_MINUTES = 10
MODEL_NAME  = "VGG-Face"
DETECTOR    = "opencv"
THRESHOLD   = 0.40
FACES_DIR   = Path("data/faces")
LOGS_DIR    = Path("attendance_logs")


# ── ATTENDANCE TRACKER ────────────────────────────────────────────────────────
class AttendanceTracker:
    def __init__(self):
        self.records         = {}
        self.lock            = threading.Lock()
        self.recent_activity = []

    def _init(self, name):
        if name not in self.records:
            self.records[name] = {"status": "absent", "last_punch_time": None}

    def process_face(self, name):
        with self.lock:
            self._init(name)
            status     = self.records[name]["status"]
            last_punch = self.records[name]["last_punch_time"]
            elapsed    = None
            if last_punch:
                elapsed = (datetime.now() - last_punch).total_seconds() / 60
            now_str = datetime.now().strftime("%H:%M:%S")

            if status == "absent":
                self.records[name]["status"]          = "signed_in"
                self.records[name]["last_punch_time"] = datetime.now()
                self._log_activity(name, "Sign-In", now_str, "")
                return "SIGNED IN", (0, 200, 0), "Sign-In", ""

            elif status == "signed_in":
                remaining = DUPLICATE_BLOCK_MINUTES - (elapsed or 0)
                if elapsed is not None and elapsed < DUPLICATE_BLOCK_MINUTES:
                    note = f"Wait {remaining:.1f} more min"
                    self._log_activity(name, "Duplicate Blocked", now_str, note)
                    return f"WAIT {remaining:.1f} min", (0, 100, 255), "Duplicate Blocked", note
                else:
                    self.records[name]["status"]          = "absent"
                    self.records[name]["last_punch_time"] = datetime.now()
                    self._log_activity(name, "Sign-Out", now_str, "")
                    return "SIGNED OUT", (255, 80, 0), "Sign-Out", ""

    def _log_activity(self, name, status, time_str, note):
        event = {"name": name, "status": status, "time": time_str, "note": note}
        self.recent_activity.insert(0, event)
        self.recent_activity = self.recent_activity[:50]

tracker = AttendanceTracker()

# ── CSV LOGGING ───────────────────────────────────────────────────────────────
def get_log_path(date_str=None):
    LOGS_DIR.mkdir(exist_ok=True)
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    return LOGS_DIR / f"attendance_{date_str}.csv"

def write_log(name, status, note=""):
    """
    Writes attendance record to CSV.    
    Logic:
        Sign-In  → creates new row with In-Time, Status=Login
        Sign-Out → finds existing row for person today, 
                   updates Out-Time, Total Hours, Status=Logout
        Duplicate → separate row just for audit (hidden in UI)
    """
    path        = get_log_path()
    file_exists = path.exists()
    now_str     = datetime.now().strftime("%H:%M:%S")

    if status == "Duplicate Blocked":
        # Write duplicate as separate audit row (hidden in UI by default)
        with open(path, "a", newline="") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Name","Date","In-Time","Out-Time","Status","Total-Hours","Note"])
            writer.writerow([name,
                             datetime.now().strftime("%Y-%m-%d"),
                             "", "", "Duplicate Blocked",
                             "", note])
        return

    # Read all existing rows
    rows = []
    if file_exists:
        with open(path, "r") as f:
            rows = list(csv.DictReader(f))

    today = datetime.now().strftime("%Y-%m-%d")

    if status == "Sign-In":
        # Add new Sign-In row for this person
        rows.append({
            "Name":         name,
            "Date":         today,
            "In-Time":      now_str,
            "Out-Time":     "",
            "Status":       "Login",
            "Total-Hours":  "",
            "Note":         ""
        })

    elif status == "Sign-Out":
        # Find the LAST Sign-In row for this person today (no out-time yet)
        updated = False
        for row in reversed(rows):
            if (row.get("Name") == name and
                row.get("Date") == today and
                row.get("Status") == "Login" and
                row.get("Out-Time") == ""):

                # Calculate total working hours
                try:
                    in_time  = datetime.strptime(f"{today} {row['In-Time']}", "%Y-%m-%d %H:%M:%S")
                    out_time = datetime.strptime(f"{today} {now_str}",        "%Y-%m-%d %H:%M:%S")
                    diff     = out_time - in_time
                    total_seconds = int(diff.total_seconds())
                    hours   = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    total_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                except:
                    total_str = ""

                row["Out-Time"]    = now_str
                row["Status"]      = "Logout"
                row["Total-Hours"] = total_str
                updated = True
                break

        if not updated:
            # No matching Sign-In found — add as standalone logout
            rows.append({
                "Name":        name,
                "Date":        today,
                "In-Time":     "",
                "Out-Time":    now_str,
                "Status":      "Logout",
                "Total-Hours": "",
                "Note":        ""
            })

    # Rewrite entire file with updated rows
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Name","Date","In-Time","Out-Time","Status","Total-Hours","Note"])
        writer.writeheader()
        writer.writerows(rows)
        
def read_log(date_str=None):
    """Read CSV for a given date. Returns list of dicts."""
    path = get_log_path(date_str)
    if not path.exists():
        return []
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        # Handle both old format and new format CSVs
        return [dict(r) for r in reader]

def read_log_range(start_str, end_str, exclude_duplicates=False):
    """Read records between start and end date inclusive."""
    records = []
    start   = datetime.strptime(start_str, "%Y-%m-%d").date()
    end     = datetime.strptime(end_str,   "%Y-%m-%d").date()
    current = start
    while current <= end:
        day_records = read_log(current.strftime("%Y-%m-%d"))
        for r in day_records:
            if exclude_duplicates and r.get("Status") == "Duplicate Blocked":
                continue
            records.append(r)
        current += timedelta(days=1)
    return records

def read_log_range(start_str, end_str, exclude_duplicates=False):
    """
    Read all CSV records between start_date and end_date (inclusive).
    Optionally filter out 'Duplicate Blocked' rows.
    """
    records = []
    start = datetime.strptime(start_str, "%Y-%m-%d").date()
    end   = datetime.strptime(end_str,   "%Y-%m-%d").date()

    current = start
    while current <= end:
        day_records = read_log(current.strftime("%Y-%m-%d"))
        for r in day_records:
            if exclude_duplicates and r.get("Status") == "Duplicate Blocked":
                continue
            records.append(r)
        current += timedelta(days=1)
    return records

def calculate_monthly_percentage(month_str):
    """
    month_str format: YYYY-MM
    Calculates attendance percentage (Mon-Sat working days)
    """
    year, month = map(int, month_str.split("-"))
    total_working_days = 0

    # Count Monday-Saturday
    month_calendar = calendar.monthcalendar(year, month)

    for week in month_calendar:
        for day_index in range(6):  # 0=Mon, 5=Sat (exclude Sunday index=6)
            if week[day_index] != 0:
                total_working_days += 1

    # Collect all logs for that month
    start = f"{month_str}-01"
    end   = f"{month_str}-{calendar.monthrange(year, month)[1]}"
    records = read_log_range(start, end, exclude_duplicates=True)

    attendance = {}

    for r in records:
        name = r.get("Name")
        if r.get("Status") == "Login":  # Count only Login as present day
            attendance.setdefault(name, set()).add(r.get("Date"))

    result = []

    for name, days_present in attendance.items():
        present_days = len(days_present)
        percentage = round((present_days / total_working_days) * 100, 2) if total_working_days else 0
        result.append({
            "Name": name,
            "Present Days": present_days,
            "Working Days": total_working_days,
            "Percentage": percentage
        })

    return result

# ── FACE DATABASE ─────────────────────────────────────────────────────────────
class FaceDatabase:
    def __init__(self):
        self.faces  = []
        self.lock   = threading.Lock()
        self.load()

    def load(self):
        new_faces = []
        FACES_DIR.mkdir(parents=True, exist_ok=True)
        for img_path in FACES_DIR.iterdir():
            if img_path.suffix.lower() not in {".jpg", ".jpeg", ".png"}:
                continue
            name = img_path.stem.replace("_", " ").replace("-", " ")
            try:
                result = DeepFace.represent(
                    img_path=str(img_path),
                    model_name=MODEL_NAME,
                    detector_backend=DETECTOR,
                    enforce_detection=True
                )
                embedding = np.array(result[0]["embedding"])
                new_faces.append({"name": name, "embedding": embedding})
                print(f"[OK] Loaded: {name}")
            except Exception as e:
                print(f"[WARN] {img_path.name}: {e}")
        with self.lock:
            self.faces = new_faces
        print(f"[INFO] {len(new_faces)} face(s) loaded.\n")

    def identify(self, frame_rgb):
        with self.lock:
            known = self.faces.copy()
        if not known:
            return []
        try:
            detections = DeepFace.represent(
                img_path=frame_rgb,
                model_name=MODEL_NAME,
                detector_backend=DETECTOR,
                enforce_detection=False
            )
        except Exception:
            return []

        results = []
        for det in detections:
            det_emb = np.array(det["embedding"])
            region  = det.get("facial_area", {})
            x, y    = region.get("x", 0), region.get("y", 0)
            w, h    = region.get("w", 0), region.get("h", 0)
            if w == 0 or h == 0:
                continue

            best_name, best_dist = "Unknown", float("inf")
            for kf in known:
                dot  = np.dot(det_emb, kf["embedding"])
                na   = np.linalg.norm(det_emb)
                nb   = np.linalg.norm(kf["embedding"])
                if na == 0 or nb == 0:
                    continue
                dist = 1 - (dot / (na * nb))
                if dist < best_dist:
                    best_dist = dist
                    best_name = kf["name"] if dist < THRESHOLD else "Unknown"
            results.append((best_name, x, y, w, h))
        return results
face_db = FaceDatabase()

# ── FRAME COOLDOWN ────────────────────────────────────────────────────────────
class FrameCooldown:
    COOLDOWN_SECONDS = 3
    def __init__(self):
        self.last = {}
    def should_process(self, name):
        if name == "Unknown": return False
        if name not in self.last: return True
        return (datetime.now() - self.last[name]).total_seconds() >= self.COOLDOWN_SECONDS
    def mark(self, name):
        self.last[name] = datetime.now()

# ── CAMERA ────────────────────────────────────────────────────────────────────
class Camera:
    def __init__(self):
        self.cap      = None
        self.frame    = None
        self.lock     = threading.Lock()
        self.running  = False
        self.cooldown = FrameCooldown()

    def start(self):
        self.cap = cv2.VideoCapture(0)
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        if not self.cap.isOpened():
            print("[ERROR] Cannot open webcam!")
            return False
        self.running = True
        threading.Thread(target=self._loop, daemon=True).start()
        print("[INFO] Camera started.")
        return True

    def stop(self):
        self.running = False
        if self.cap: self.cap.release()

    def get_frame(self):
        with self.lock:
            return self.frame

    def _loop(self):
        frame_count  = 0
        display_data = {}
        while self.running:
            ret, frame = self.cap.read()
            if not ret:
                time.sleep(0.05)
                continue
            frame_count += 1

            if frame_count % 5 == 0:
                rgb          = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                detections   = face_db.identify(rgb)
                current_names = set()

                for name, x, y, w, h in detections:
                    current_names.add(name)
                    if name == "Unknown":
                        action, color = "NOT REGISTERED", (0, 0, 200)
                    else:
                        if self.cooldown.should_process(name):
                            action, color, status, note = tracker.process_face(name)
                            self.cooldown.mark(name)
                            write_log(name, status, note)
                        else:
                            if name in display_data:
                                _, _, _, _, action, color = display_data[name]
                            else:
                                action, color = "", (0, 200, 0)
                    display_data[name] = (x, y, w, h, action, color)

                for n in list(display_data.keys()):
                    if n not in current_names:
                        del display_data[n]

            for name, (x, y, w, h, action, color) in display_data.items():
                cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                cv2.rectangle(frame, (x, y+h), (x+w, y+h+55), color, cv2.FILLED)
                cv2.putText(frame, name,   (x+6, y+h+22), cv2.FONT_HERSHEY_DUPLEX, 0.6,  (255,255,255), 1)
                cv2.putText(frame, action, (x+6, y+h+48), cv2.FONT_HERSHEY_DUPLEX, 0.42, (255,255,255), 1)

            overlay = frame.copy()
            cv2.rectangle(overlay, (0,0), (frame.shape[1], 55), (0,0,0), cv2.FILLED)
            cv2.addWeighted(overlay, 0.55, frame, 0.45, 0, frame)
            cv2.putText(frame, "FACE RECOGNITION ATTENDANCE  |  DeepFace Engine",
                        (10, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 1)

            _, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            with self.lock:
                self.frame = jpeg.tobytes()
camera = Camera()

# ── AUTH HELPERS ──────────────────────────────────────────────────────────────
def is_admin():
    """Returns True if current session is an authenticated admin."""
    return session.get("admin_logged_in") is True

# ── FLASK ROUTES ──────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html", is_admin=is_admin())

@app.route("/video_feed")
def video_feed():
    def generate():
        while True:
            frame = camera.get_frame()
            if frame is None:
                time.sleep(0.05)
                continue
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + frame + b"\r\n")
            time.sleep(0.033)
    return Response(generate(), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/api/monthly_percentage")
def monthly_percentage():
    month = request.args.get("month")  # format: YYYY-MM
    if not month:
        return jsonify({"error": "Month required"}), 400

    data = calculate_monthly_percentage(month)
    return jsonify({"records": data})


# ── ADMIN AUTH ROUTES ──────────────────────────────────────────────────────────
@app.route("/admin/login", methods=["POST"])
def admin_login():
    """
    Verifies admin credentials.
    Request JSON: { "username": "admin", "password": "admin123" }
    Sets session cookie on success.
    """
    data     = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session["admin_logged_in"] = True
        return jsonify({"success": True,  "message": "Login successful!"})
    else:
        return jsonify({"success": False, "message": "Invalid username or password."})

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    """Clears the admin session."""
    session.pop("admin_logged_in", None)
    return jsonify({"success": True})

@app.route("/admin/status")
def admin_status():
    """Returns whether admin is currently logged in."""
    return jsonify({"is_admin": is_admin()})

# ── REGISTER FACE (admin only) ────────────────────────────────────────────────
@app.route("/api/register", methods=["POST"])
def register_face():
    """
    Registers a new face. ADMIN ONLY.
    Returns 403 if not logged in as admin.
    """
    if not is_admin():
        return jsonify({"success": False, "message": "Unauthorized. Admin login required."}), 403

    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "message": "Name cannot be empty."})

    frame_bytes = camera.get_frame()
    if frame_bytes is None:
        return jsonify({"success": False, "message": "Camera not ready."})

    nparr = np.frombuffer(frame_bytes, np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    try:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        DeepFace.represent(img_path=rgb, model_name=MODEL_NAME,
                           detector_backend=DETECTOR, enforce_detection=True)
    except Exception:
        return jsonify({"success": False, "message": "No face detected! Look at camera."})

    filename  = name.replace(" ", "_") + ".jpg"
    save_path = FACES_DIR / filename
    cv2.imwrite(str(save_path), frame)
    face_db.load()

    return jsonify({"success": True, "message": f"'{name}' registered successfully!"})

# ── ATTENDANCE API ────────────────────────────────────────────────────────────
@app.route("/api/activity")
def get_activity():
    with tracker.lock:
        events = tracker.recent_activity[:20]
    return jsonify({"events": events})

@app.route("/api/attendance")
def get_attendance():
    """
    Returns attendance records for today.
    Query param: ?exclude_duplicates=true (hides duplicate blocked rows)
    """
    date_str           = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    exclude_duplicates = request.args.get("exclude_duplicates", "true").lower() == "true"
    records            = read_log(date_str)

    if exclude_duplicates:
        records = [r for r in records if r.get("Status") != "Duplicate Blocked"]

    return jsonify({"records": records, "date": date_str})

@app.route("/api/registered")
def get_registered():
    with face_db.lock:
        people = [f["name"] for f in face_db.faces]
    return jsonify({"people": people})

@app.route("/api/dates")
def get_dates():
    LOGS_DIR.mkdir(exist_ok=True)
    dates = []
    for f in sorted(LOGS_DIR.iterdir(), reverse=True):
        if f.name.startswith("attendance_") and f.name.endswith(".csv"):
            dates.append(f.stem.replace("attendance_", ""))
    return jsonify({"dates": dates})

# ── DOWNLOAD ROUTES ───────────────────────────────────────────────────────────
@app.route("/api/download/excel")
def download_excel():
    """
    Downloads attendance records as Excel (.xlsx) for a date range.
    Query params: ?start=YYYY-MM-DD&end=YYYY-MM-DD&exclude_duplicates=true

    Uses openpyxl to build a styled spreadsheet in memory,
    then sends it as a file download.
    """
    start              = request.args.get("start", datetime.now().strftime("%Y-%m-%d"))
    end                = request.args.get("end",   datetime.now().strftime("%Y-%m-%d"))
    exclude_duplicates = request.args.get("exclude_duplicates", "true").lower() == "true"

    records = read_log_range(start, end, exclude_duplicates)

    # Build Excel workbook in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value     = f"Attendance Report  |  {start}  to  {end}"
    title_cell.font      = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1a73e8")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # ── Header row ─────────────────────────────────────────────────────────
    headers = ["Name", "Date", "In-Time", "Out-Time", "Status", "Total Hours"]
    header_row = 2
    for col, header in enumerate(headers, 1):
        cell            = ws.cell(row=header_row, column=col, value=header)
        cell.font       = Font(bold=True, color="FFFFFF")
        cell.fill       = PatternFill("solid", fgColor="333333")
        cell.alignment  = Alignment(horizontal="center")

    # ── Data rows ──────────────────────────────────────────────────────────
    status_colors = {
        "Sign-In":  "d4edda",   # Light green
        "Sign-Out": "cce5ff",   # Light blue
        "Duplicate Blocked": "fff3cd"  # Light orange
    }

    for row_idx, record in enumerate(records, start=3):
        ws.cell(row=row_idx, column=1, value=record.get("Name",        ""))
        ws.cell(row=row_idx, column=2, value=record.get("Date",        ""))
        ws.cell(row=row_idx, column=3, value=record.get("In-Time",     ""))
        ws.cell(row=row_idx, column=4, value=record.get("Out-Time",    ""))
        ws.cell(row=row_idx, column=5, value=record.get("Status",      ""))
        ws.cell(row=row_idx, column=6, value=record.get("Total-Hours", ""))

        status     = record.get("Status", "")
        fill_color = status_colors.get(status, "ffffff")
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=fill_color)

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_{start}_to_{end}.xlsx"
    return send_file(buffer, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/download/pdf")
def download_pdf():
    """
    Downloads attendance records as PDF for a date range.
    Query params: ?start=YYYY-MM-DD&end=YYYY-MM-DD&exclude_duplicates=true

    Uses ReportLab to build a styled PDF table in memory.
    """
    start              = request.args.get("start", datetime.now().strftime("%Y-%m-%d"))
    end                = request.args.get("end",   datetime.now().strftime("%Y-%m-%d"))
    exclude_duplicates = request.args.get("exclude_duplicates", "true").lower() == "true"

    records = read_log_range(start, end, exclude_duplicates)

    # Build PDF in memory
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=40, rightMargin=40,
                                topMargin=40,  bottomMargin=40)
    styles  = getSampleStyleSheet()
    content = []

    # Title
    title = Paragraph(
        f"<b>Attendance Report</b><br/><font size=10>{start} to {end}</font>",
        styles["Title"]
    )
    content.append(title)
    content.append(Spacer(1, 20))

    # Table data
    table_data = [["Name", "Date", "In-Time", "Out-Time", "Status", "Total Hours"]]  # Header row
    for r in records:
        table_data.append([
            r.get("Name",        ""),
            r.get("Date",        ""),
            r.get("In-Time",     ""),
            r.get("Out-Time",    ""),
            r.get("Status",      ""),
            r.get("Total-Hours", "")
        ])

    # Build table
    table = Table(table_data, colWidths=[110, 75, 60, 60, 70, 80])

    # Table styles
    table_style = TableStyle([
        # Header row styling
        ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#1a73e8")),
        ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
        ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0),  11),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUND",(0,1),(-1,-1), [colors.HexColor("#f8f9fa"), colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#dee2e6")),
        ("ROWHEIGHT",   (0,0), (-1,-1), 24),
        ("FONTSIZE",    (0,1), (-1,-1), 9),
    ])

    # Color data rows by status
    for row_idx, record in enumerate(records, start=1):
        status = record.get("Status", "")
        if status == "Sign-In":
            table_style.add("BACKGROUND", (0,row_idx), (-1,row_idx), colors.HexColor("#d4edda"))
        elif status == "Sign-Out":
            table_style.add("BACKGROUND", (0,row_idx), (-1,row_idx), colors.HexColor("#cce5ff"))
        elif status == "Duplicate Blocked":
            table_style.add("BACKGROUND", (0,row_idx), (-1,row_idx), colors.HexColor("#fff3cd"))

    table.setStyle(table_style)
    content.append(table)

    # Summary
    content.append(Spacer(1, 20))
    sign_ins  = sum(1 for r in records if r.get("Status") == "Sign-In")
    sign_outs = sum(1 for r in records if r.get("Status") == "Sign-Out")
    summary   = Paragraph(
        f"Total Records: {len(records)}  &nbsp;|&nbsp;  Sign-Ins: {sign_ins}  &nbsp;|&nbsp;  Sign-Outs: {sign_outs}",
        styles["Normal"]
    )
    content.append(summary)

    doc.build(content)
    buffer.seek(0)

    filename = f"attendance_{start}_to_{end}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename,
                     mimetype="application/pdf")

# ── STARTUP ───────────────────────────────────────────────────────────────────
def restore_tracker_state():
    """
    On startup, rebuild tracker state from today's CSV.
    Now reads new format: Login/Logout instead of Sign-In/Sign-Out
    """
    records = read_log()
    if not records:
        print("[INFO] No existing records today. Fresh start.")
        return

    print("[INFO] Restoring attendance state from today's CSV...")

    for record in records:
        name   = record.get("Name",   "").strip()
        status = record.get("Status", "").strip()

        if not name or status == "Duplicate Blocked":
            continue

        if name not in tracker.records:
            tracker.records[name] = {"status": "absent", "last_punch_time": None}

        today = datetime.now().strftime("%Y-%m-%d")

        if status == "Login":
            tracker.records[name]["status"] = "signed_in"
            try:
                t = record.get("In-Time", "")
                tracker.records[name]["last_punch_time"] = datetime.strptime(
                    f"{today} {t}", "%Y-%m-%d %H:%M:%S")
            except:
                tracker.records[name]["last_punch_time"] = datetime.now()

        elif status == "Logout":
            tracker.records[name]["status"] = "absent"
            try:
                t = record.get("Out-Time", "")
                tracker.records[name]["last_punch_time"] = datetime.strptime(
                    f"{today} {t}", "%Y-%m-%d %H:%M:%S")
            except:
                tracker.records[name]["last_punch_time"] = datetime.now()

    for name, state in tracker.records.items():
        print(f"  [RESTORED] {name:20s} → {state['status']}")
    print()

if __name__ == "__main__":
    print("=" * 60)
    print("  FACE ATTENDANCE SYSTEM v2 — DeepFace + Flask")
    print("=" * 60)

    FACES_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(exist_ok=True)

    #restores state from today's CSV
    restore_tracker_state()

    print("\n[INFO] Starting camera...")
    camera.start()
    print("[INFO] Open browser: http://localhost:5000\n")
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)